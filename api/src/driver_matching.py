"""
Fuzzy-match driver roster names against an SSN export (real callout PINs)
and a Slack workspace member export (Slack IDs for driver DMs).

Extracted 2026-07-15 from scripts/import_ssn_slack.py so the same logic
is usable both as a local CLI script and as a proper API upload endpoint
(api/src/routes/drivers.py) — the app never needs raw production DB
credentials handed to a local script when this exists.
"""
from __future__ import annotations

import csv
import os
import re
from difflib import SequenceMatcher

MATCH_THRESHOLD = 0.82            # SequenceMatcher ratio cutoff
ASSOCIATE_MATCH_THRESHOLD = 0.90  # stricter — feeds a deterministic email lookup, a wrong hit here silently links the wrong Slack account


def _read_rows(path: str) -> list[dict]:
    """Read a header-row table from either .xlsx or .csv into a list of
    dicts keyed by header name. CSV support added 2026-07-16 for the
    SlackMemberExtractor Chrome-extension export, which ships CSV with a
    UTF-8 BOM rather than xlsx."""
    if os.path.splitext(path)[1].lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdrs = [str(c.value).strip() if c.value else "" for c in ws[1]]
    return [dict(zip(hdrs, raw)) for raw in ws.iter_rows(min_row=2, values_only=True)]


def _norm(s: str) -> str:
    """Lowercase, collapse whitespace, strip punctuation for fuzzy compare."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", "", s.lower())).strip()


def _ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


def _build_ssn_candidates(first: str, middle: str, last: str) -> list[str]:
    """All plausible full-name formats from SSN file columns."""
    first  = (first  or "").strip()
    middle = (middle or "").strip()
    last   = (last   or "").strip()
    candidates = []
    if first and middle and last:
        candidates.append(f"{first} {middle} {last}")
        candidates.append(f"{first} {last}")
        candidates.append(f"{last} {first} {middle}")
        candidates.append(f"{last} {first}")
    elif first and last:
        candidates.append(f"{first} {last}")
        candidates.append(f"{last} {first}")
    return candidates


def load_ssn(path: str) -> list[dict]:
    """Load SSN export (xlsx) → list of {candidates: [...], last4}"""
    rows = []
    for row in _read_rows(path):
        last4 = str(row.get("last 4") or "").strip().zfill(4)
        first  = str(row.get("Legal First Name")   or "").strip()
        middle = str(row.get("Legal Middle Name")  or "").strip()
        last   = str(row.get("Legal Last Name")    or "").strip()
        if not last4 or not (first or last):
            continue
        rows.append({"candidates": _build_ssn_candidates(first, middle, last), "last4": last4})
    return rows


def load_slack(path: str) -> list[dict]:
    """Load a Slack workspace member export (xlsx or CSV) → list of
    {user_id, username, display_name}. Supports two header conventions:
    the original xlsx export ("User ID"/"Username"/"Display Name") and the
    SlackMemberExtractor CSV tool ("Id"/"Name"/"Real Name" — CSV's
    "Display Name" is often blank, so "Real Name" is the fallback)."""
    rows = []
    for row in _read_rows(path):
        user_id      = str(row.get("User ID") or row.get("Id") or "").strip()
        username     = str(row.get("Username") or row.get("Name") or "").strip()
        display_name = (
            str(row.get("Display Name") or "").strip()
            or str(row.get("Real Name") or "").strip()
        )
        if user_id:
            rows.append({"user_id": user_id, "username": username, "display_name": display_name})
    return rows


def load_associates(path: str) -> list[dict]:
    """Load an Amazon associate/Transporter roster export (xlsx or CSV,
    e.g. "AssociateData (N).csv") → list of {name, email}. Bridges a
    roster driver's legal name to a Slack account via email local-part —
    far more reliable than fuzzy-matching directly against Slack's often
    auto-generated "Real Name" field (e.g. "A Laporte Ndl", generated
    from the username, not a real display name someone typed in)."""
    rows = []
    for row in _read_rows(path):
        name = str(row.get("Name and ID") or row.get("Name") or "").strip()
        email = str(row.get("Email") or "").strip()
        if name and email and "@" in email:
            rows.append({"name": name, "email": email})
    return rows


def best_slack_match_via_associates(
    roster_name: str, associate_rows: list[dict], slack_rows: list[dict]
) -> tuple[str | None, str | None, float]:
    """Bridge roster_name -> associate legal name (strict fuzzy match) ->
    email local-part -> exact match against the Slack export's username
    column. Returns (None, None, best_associate_score) if any link in the
    chain fails — the caller should fall back to best_slack_match()
    (direct fuzzy match against Slack display/real names) in that case."""
    slack_by_username = {
        r["username"].strip().lower(): r for r in slack_rows if r["username"]
    }

    best_assoc, best_score = None, 0.0
    for a in associate_rows:
        score = _ratio(roster_name, a["name"])
        if score > best_score:
            best_score = score
            best_assoc = a
    if best_score < ASSOCIATE_MATCH_THRESHOLD or not best_assoc:
        return None, None, best_score

    local_part = best_assoc["email"].split("@")[0].strip().lower()
    slack_row = slack_by_username.get(local_part)
    if not slack_row:
        return None, None, best_score

    return slack_row["user_id"], (slack_row["display_name"] or slack_row["username"]), best_score


def best_ssn_match(roster_name: str, ssn_rows: list[dict]) -> tuple[str | None, float]:
    """Find the SSN row whose candidate names best match roster_name."""
    best_last4, best_score = None, 0.0
    for row in ssn_rows:
        for cand in row["candidates"]:
            score = _ratio(roster_name, cand)
            if score > best_score:
                best_score = score
                best_last4 = row["last4"]
    if best_score >= MATCH_THRESHOLD:
        return best_last4, best_score
    return None, best_score


def best_slack_match(roster_name: str, slack_rows: list[dict]) -> tuple[str | None, str | None, float]:
    """Find Slack user whose display_name or username best matches roster_name."""
    best_id, best_display, best_score = None, None, 0.0
    parts = _norm(roster_name).split()
    first = parts[0] if parts else ""
    last  = parts[-1] if len(parts) > 1 else ""

    for row in slack_rows:
        candidates = []
        if row["display_name"]:
            candidates.append(row["display_name"])
        if row["username"]:
            candidates.append(row["username"])
        for cand in candidates:
            score = _ratio(roster_name, cand)
            # Boost if first OR last name appears in the candidate
            nc = _norm(cand)
            if first in nc or last in nc:
                score = max(score, 0.60)
            if score > best_score:
                best_score = score
                best_id      = row["user_id"]
                best_display = row["display_name"] or row["username"]

    if best_score >= MATCH_THRESHOLD:
        return best_id, best_display, best_score
    return None, None, best_score
