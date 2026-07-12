"""Parse the Amazon DVIC Pre-Trip Under-90-Second weekly Excel report."""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Optional

import openpyxl


def _parse_date(val) -> Optional[date]:
    if not val:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(val) -> Optional[datetime]:
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def extract_week(filename: str) -> str:
    """Return '2026-W27' from a filename like ..._week-27_..."""
    m = re.search(r'week[-_](\d+)', filename, re.IGNORECASE)
    if m:
        year_m = re.search(r'(20\d{2})', filename)
        year = year_m.group(1) if year_m else str(datetime.utcnow().year)
        return f"{year}-W{int(m.group(1)):02d}"
    return "unknown"


_KNOWN_HEADERS = {
    "start_date", "dsp", "station", "transporter_id", "transporter_name",
    "vin", "fleet_type", "inspection_type", "inspection_status",
    "start_time", "end_time", "duration",
}


def parse_dvic_xlsx(content: bytes, filename: str) -> tuple[dict, list[dict]]:
    """
    Parse a DVIC Pre-Trip under-90s Excel file.

    Returns:
        summary dict: week, total_violations, unique_drivers, date_range_start, date_range_end
        violations list: one dict per row
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    sheet_name = "DVIC Detail" if "DVIC Detail" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"week": "unknown", "total_violations": 0, "unique_drivers": 0,
                "date_range_start": None, "date_range_end": None}, []

    headers = [str(h).lower().strip() if h else "" for h in rows[0]]

    violations: list[dict] = []
    for raw in rows[1:]:
        row = dict(zip(headers, raw))
        tid = str(row.get("transporter_id") or "").strip()
        if not tid:
            continue

        dur = row.get("duration")
        try:
            dur_int = int(dur) if dur is not None else None
        except (ValueError, TypeError):
            dur_int = None

        # Governed by the file's own headers: anything the sheet includes
        # beyond the known/mapped columns is kept verbatim rather than
        # silently dropped, so future report variants don't lose data.
        raw_fields = {}
        for key, val in row.items():
            if not key or key in _KNOWN_HEADERS or val is None:
                continue
            raw_fields[key] = val.isoformat() if isinstance(val, (datetime, date)) else str(val)

        violations.append({
            "start_date":        _parse_date(row.get("start_date")),
            "dsp":               str(row.get("dsp") or "").strip() or None,
            "station":           str(row.get("station") or "").strip() or None,
            "transporter_id":    tid,
            "transporter_name":  str(row.get("transporter_name") or "").strip() or None,
            "vin":               str(row.get("vin") or "").strip() or None,
            "fleet_type":        str(row.get("fleet_type") or "").strip() or None,
            "inspection_type":   str(row.get("inspection_type") or "").strip() or None,
            "inspection_status": str(row.get("inspection_status") or "").strip() or None,
            "start_time":        _parse_datetime(row.get("start_time")),
            "end_time":          _parse_datetime(row.get("end_time")),
            "duration_seconds":  dur_int,
            "raw_fields":        raw_fields or None,
        })

    week = extract_week(filename)
    start_dates = [v["start_date"] for v in violations if v["start_date"]]
    unique_drivers = len(set(v["transporter_id"] for v in violations))

    summary = {
        "week":              week,
        "total_violations":  len(violations),
        "unique_drivers":    unique_drivers,
        "date_range_start":  str(min(start_dates)) if start_dates else None,
        "date_range_end":    str(max(start_dates)) if start_dates else None,
    }
    return summary, violations
