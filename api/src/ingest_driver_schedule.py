"""Driver schedule ingest from Excel file (Rostered Work Blocks and Shifts & Availability tabs)."""
import re
from typing import List, Dict, Tuple, Optional
from datetime import datetime, timedelta
import openpyxl
from api.src.models import DriverAssignment, DriverAvailability, DriverScheduleSummary


def parse_driver_schedule_excel(file_path: str) -> Tuple[DriverScheduleSummary, List[str]]:
    """
    Parse driver schedule Excel file with two tabs:
    1. 'Rostered Work Blocks': Driver assignments for night prior (row 4 = dates, col A = names)
    2. 'Shifts & Availability': Driver availability (same structure)
    
    Returns:
        DriverScheduleSummary with show times and sweeper list
        List of error/warning messages
    """
    errors = []
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Get timestamp from A2 (should be same on both tabs)
        timestamp_str = ""
        if 'Rostered Work Blocks' in wb.sheetnames:
            ws_assigned = wb['Rostered Work Blocks']
            timestamp_cell = ws_assigned['A2'].value
            if timestamp_cell:
                timestamp_str = str(timestamp_cell).strip()
        
        if not timestamp_str:
            errors.append("Could not extract timestamp from A2. File may be invalid.")
            return DriverScheduleSummary(timestamp="", date=""), errors
        
        # Parse Rostered Work Blocks tab
        if 'Rostered Work Blocks' not in wb.sheetnames:
            errors.append("'Rostered Work Blocks' tab not found in file")
            return DriverScheduleSummary(timestamp=timestamp_str, date=""), errors
        
        ws_assigned = wb['Rostered Work Blocks']
        assigned_drivers, assigned_dates, assignments = _parse_work_blocks_tab(ws_assigned, errors)
        
        # Parse Shifts & Availability tab
        availability_by_driver = {}
        if 'Shifts & Availability' in wb.sheetnames:
            ws_availability = wb['Shifts & Availability']
            availability_by_driver = _parse_availability_tab(ws_availability, errors)
        else:
            errors.append("'Shifts & Availability' tab not found in file")
        
        # Determine the scheduled date (first date from row 4)
        scheduled_date = assigned_dates[0] if assigned_dates else ""
        
        # Filter assignments to only those matching the scheduled date
        assignments_for_scheduled_date = [a for a in assignments if a.date == scheduled_date]
        
        # Calculate show times from assignments
        assignments_with_show_times = _calculate_show_times(assignments_for_scheduled_date)
        
        # Find sweepers (scheduled but not assigned)
        # A driver is a sweeper if:
        # 1. They are scheduled (have availability) on Tab 2 for the scheduled date
        # 2. The availability status is NOT "Unavailable" and NOT blank
        # 3. They are NOT in the assignments list for the scheduled date
        sweepers = _identify_sweepers(
            set(driver.driver_name for driver in assignments_with_show_times),
            availability_by_driver,
            scheduled_date
        )
        
        # Get earliest wave show time for sweepers
        earliest_show_time = _get_earliest_show_time(assignments_with_show_times)
        
        # Build summary
        summary = DriverScheduleSummary(
            timestamp=timestamp_str,
            date=scheduled_date,
            assignments=assignments_with_show_times,
            sweepers=sweepers,
        )
        
        # Build show_times dictionary
        show_times_dict = {}
        for assignment in assignments_with_show_times:
            show_times_dict[assignment.driver_name] = assignment.show_time or ""
        
        # Add sweepers with earliest show time
        for sweeper in sweepers:
            show_times_dict[sweeper] = earliest_show_time
        
        summary.show_times = show_times_dict
        
        return summary, errors
    
    except Exception as e:
        errors.append(f"Failed to parse driver schedule Excel: {str(e)}")
        return DriverScheduleSummary(timestamp="", date=""), errors


def _parse_work_blocks_tab(ws, errors: List[str]) -> Tuple[List[str], List[str], List[DriverAssignment]]:
    """
    Parse 'Rostered Work Blocks' tab.
    Row 4 contains dates starting from Column C.
    Column A contains driver names starting from Row 6.
    Assignment data starts in Column C, Row 6 onwards.
    
    Returns:
        Tuple of (driver_names, dates, assignments)
    """
    assignments = []
    driver_names = []
    dates = []
    
    try:
        # Extract dates from row 4 (starting from column C, which is column 3)
        col_idx = 3  # Start from column C
        for col_offset in range(10):  # Check up to 10 columns
            cell = ws.cell(row=4, column=col_idx + col_offset)
            if cell.value:
                date_str = str(cell.value).strip()
                # Only add if it looks like a date (contains day abbreviation)
                if any(day in date_str for day in ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']):
                    dates.append(date_str)
                else:
                    break  # Stop when we hit non-date columns
        
        # Extract driver names from column A (starting from row 6)
        for row_idx in range(6, ws.max_row + 1):
            driver_cell = ws.cell(row=row_idx, column=1)
            if not driver_cell.value:
                break  # Stop at empty rows
            
            driver_name = str(driver_cell.value).strip()
            # Skip "Total" rows
            if "Total" not in driver_name and driver_name:
                driver_names.append(driver_name)
        
        # Extract assignments: intersection of driver and date
        for row_idx, driver_name in enumerate(driver_names, start=6):  # Row 6 onwards
            for col_idx, date in enumerate(dates, start=3):  # Column C onwards (column 3)
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    assignment_data = str(cell.value).strip()
                    # Parse assignment data  
                    wave_time, service_type = _parse_assignment_data(assignment_data)
                    
                    if wave_time:  # Only create assignment if wave time found
                        assignment = DriverAssignment(
                            driver_name=driver_name,
                            date=date,
                            wave_time=wave_time,
                            service_type=service_type,
                        )
                        assignments.append(assignment)
        
        return driver_names, dates, assignments
    
    except Exception as e:
        errors.append(f"Error parsing Rostered Work Blocks tab: {str(e)}")
        return [], [], []


def _parse_availability_tab(ws, errors: List[str]) -> Dict[str, List[DriverAvailability]]:
    """
    Parse 'Shifts & Availability' tab.
    Row 4 contains dates starting from Column C.
    Column A contains driver names starting from Row 6.
    
    Returns:
        Dictionary mapping driver names to list of DriverAvailability records
    """
    availability_by_driver = {}
    
    try:
        # Extract dates from row 4 (starting from column C, which is column 3)
        dates = []
        for col_offset in range(10):  # Check up to 10 columns
            cell = ws.cell(row=4, column=3 + col_offset)
            if cell.value:
                date_str = str(cell.value).strip()
                # Only add if it looks like a date
                if any(day in date_str for day in ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']):
                    dates.append(date_str)
                else:
                    break
        
        # Extract driver-availability pairs (starting from row 6, column A)
        for row_idx in range(6, ws.max_row + 1):
            driver_cell = ws.cell(row=row_idx, column=1)
            if not driver_cell.value:
                break  # Stop at empty rows
            
            driver_name = str(driver_cell.value).strip()
            if "Total" in driver_name or not driver_name:
                break
            
            availabilities = []
            for col_idx, date in enumerate(dates, start=3):  # Column C onwards (column 3)
                av_cell = ws.cell(row=row_idx, column=col_idx)
                if av_cell.value:
                    av_data = str(av_cell.value).strip()
                    av = DriverAvailability(
                        driver_name=driver_name,
                        date=date,
                        availability=av_data,
                    )
                    availabilities.append(av)
            
            if availabilities:
                availability_by_driver[driver_name] = availabilities
        
        return availability_by_driver
    
    except Exception as e:
        errors.append(f"Error parsing Shifts & Availability tab: {str(e)}")
        return {}


def _parse_assignment_data(data_str: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Parse assignment data string.
    Format can be:
    - "10:20am • 10 hrs"
    - "10:20 AM • 10 hrs"
    - Multi-line: "Standard Parcel Electric - Rivian MEDIUM\n10:20am • 10 hrs"
    - "Group 1" (from availability tab, not a valid assignment)
    
    Returns:
        Tuple of (wave_time, service_type)
    """
    # Look for time pattern like "10:20 AM" or "10:20am"
    time_pattern = r'(\d{1,2}):(\d{2})\s*(?:AM|PM|am|pm)'
    time_match = re.search(time_pattern, data_str, re.IGNORECASE)
    
    if not time_match:
        return None, None
    
    # Extract hour and minute
    hour = int(time_match.group(1))
    minute = int(time_match.group(2))
    
    # Determine AM/PM by looking at the match
    full_match = data_str[time_match.start():time_match.end()]
    am_pm = "AM" if full_match.upper().endswith("AM") else "PM"
    if full_match.upper().endswith("PM"):
        am_pm = "PM"
    
    # Format as "H:MM AM/PM"
    wave_time = f"{hour}:{minute:02d} {am_pm}"
    
    # Extract service type (before the time, usually on first line)
    service_type = None
    lines = data_str.split('\n')
    for line in lines:
        line = line.strip()
        if 'Electric' in line or 'Truck' in line or 'Van' in line or 'Parcel' in line:
            service_type = line
            break
    
    return wave_time, service_type


def _calculate_show_times(assignments: List[DriverAssignment]) -> List[DriverAssignment]:
    """
    Calculate show times (25 minutes before wave time).
    Wave consolidation: If waves are 5 minutes apart, use same show time.
    - Wave 1 & 2 same show time
    - Wave 3 & 4 same show time
    
    Returns:
        Updated assignments with show_time calculated
    """
    # First pass: Extract unique wave times and calculate show times
    wave_times = set()
    for assignment in assignments:
        if assignment.wave_time:
            wave_times.add(assignment.wave_time)
    
    # Build mapping of wave_time -> show_time
    wave_to_show_time = {}
    sorted_waves = sorted(wave_times, key=_time_to_minutes)
    
    grouped_waves = []
    current_group = []
    
    for i, wave in enumerate(sorted_waves):
        if not current_group:
            current_group.append(wave)
        else:
            # Check if this wave is within 5 minutes of the last one in the group
            prev_wave_minutes = _time_to_minutes(current_group[-1])
            curr_wave_minutes = _time_to_minutes(wave)
            if abs(curr_wave_minutes - prev_wave_minutes) <= 5:
                current_group.append(wave)
            else:
                # Start new group
                grouped_waves.append(current_group)
                current_group = [wave]
    
    if current_group:
        grouped_waves.append(current_group)
    
    # Calculate show time for each group (25 min before earliest wave in group)
    for group in grouped_waves:
        earliest_wave = min(group, key=_time_to_minutes)
        show_time = _subtract_minutes_from_time(earliest_wave, 25)
        for wave in group:
            wave_to_show_time[wave] = show_time
    
    # Apply show times to assignments
    for assignment in assignments:
        if assignment.wave_time in wave_to_show_time:
            assignment.show_time = wave_to_show_time[assignment.wave_time]
    
    return assignments


def _time_to_minutes(time_str: str) -> int:
    """Convert time string (e.g., '10:20 AM') to minutes since midnight."""
    try:
        # Parse time string
        time_match = re.match(r'(\d{1,2}):(\d{2})\s*(AM|PM)', time_str, re.IGNORECASE)
        if not time_match:
            return 0
        
        hour = int(time_match.group(1))
        minute = int(time_match.group(2))
        am_pm = time_match.group(3).upper()
        
        # Convert to 24-hour format
        if am_pm == 'PM' and hour != 12:
            hour += 12
        elif am_pm == 'AM' and hour == 12:
            hour = 0
        
        return hour * 60 + minute
    except:
        return 0


def _subtract_minutes_from_time(time_str: str, minutes: int) -> str:
    """Subtract minutes from a time string."""
    try:
        total_minutes = _time_to_minutes(time_str)
        new_minutes = max(0, total_minutes - minutes)
        
        hour = new_minutes // 60
        minute = new_minutes % 60
        
        # Convert back to 12-hour format
        am_pm = 'AM'
        if hour >= 12:
            am_pm = 'PM'
            if hour > 12:
                hour -= 12
        elif hour == 0:
            hour = 12
        
        return f"{hour}:{minute:02d} {am_pm}"
    except:
        return time_str


def _identify_sweepers(
    assigned_driver_names: set,
    availability_by_driver: Dict[str, List[DriverAvailability]],
    scheduled_date: str
) -> List[str]:
    """
    Identify sweepers: drivers scheduled (in Shifts & Availability) but not assigned (in Rostered Work Blocks).
    A driver is a sweeper if:
    - They are scheduled on Tab 2 for the scheduled_date
    - The availability value is NOT "Unavailable" and NOT blank
    - They are NOT in the assignments list for the scheduled_date
    """
    sweepers = []
    
    for driver_name, availabilities in availability_by_driver.items():
        # Only consider if driver is NOT already assigned with a wave time
        if driver_name not in assigned_driver_names:
            # Check if driver has availability on the SCHEDULED DATE
            # Valid availability: anything except "Unavailable" or blank
            for av in availabilities:
                if av.date == scheduled_date:
                    availability_status = av.availability.lower().strip()
                    # Include if not unavailable and not blank
                    if availability_status and availability_status != 'unavailable':
                        sweepers.append(driver_name)
                        break  # Found valid availability for this driver on this date
    
    return sorted(sweepers)


def _get_earliest_show_time(assignments: List[DriverAssignment]) -> str:
    """Get the earliest show time from assignments (Wave 1 & 2 show time)."""
    show_times = [a.show_time for a in assignments if a.show_time]
    if not show_times:
        return ""
    
    earliest = min(show_times, key=_time_to_minutes)
    return earliest
