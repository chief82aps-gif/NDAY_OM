"""
Ops Channel Ingest — watch #nday-operations-management for file uploads.

Workflow:
  1. Background loop (every 60 s) scans the channel via conversations.history.
  2. Every new file share creates an OpsIngestJob row (status=pending).
  3. Admin opens /ops-ingest, sees the queue, reads the description, clicks Ingest.
  4. The ingest endpoint downloads the file from Slack and dispatches to the
     correct handler based on detected_type.
  5. After success, a Slack confirmation is posted back to the channel.

File-type handlers are added one at a time as new file types are introduced.
"""
from __future__ import annotations

import io
import logging
import os
import re
import tempfile
from datetime import datetime, timezone
from typing import Optional

import requests
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session

from api.src.database import get_db, OpsIngestJob

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/ops-ingest", tags=["ops-ingest"])

OPS_CHANNEL = os.getenv("OPS_CHANNEL_ID", "C0BE4ALL1EX")  # #nday-operations-management
DLV3_INFO_CHANNEL = os.getenv("DLV3_INFO_CHANNEL_ID", "C0AF48TPAMV")  # #dlv3-nday-info (DOP + route sheets)
SCAN_CHANNELS = [OPS_CHANNEL, DLV3_INFO_CHANNEL]


# ─────────────────────────────────────────────────────────────────────────────
# Slack helpers
# ─────────────────────────────────────────────────────────────────────────────

def _slack_client():
    token = os.getenv("SLACK_BOT_TOKEN")
    if not token:
        return None
    from slack_sdk import WebClient
    return WebClient(token=token)


def _download_file(url: str) -> Optional[bytes]:
    token = os.getenv("SLACK_BOT_TOKEN")
    if not token:
        return None
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
        resp.raise_for_status()
        return resp.content
    except Exception as exc:
        logger.warning("Ops ingest download failed: %s", exc)
        return None


def _post_confirmation(text: str) -> None:
    client = _slack_client()
    if not client:
        return
    try:
        client.chat_postMessage(channel=OPS_CHANNEL, text=text)
    except Exception as exc:
        logger.warning("Ops ingest Slack confirm failed: %s", exc)


def _parse_all_schedule_dates(file_path: str) -> dict:
    """
    Read the 'Shifts & Availability' tab and return {date_obj: [driver_name, ...]}
    for every date column in the file. Excludes blank/Unavailable cells.
    """
    from datetime import date as _date, datetime as _dt
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if 'Shifts & Availability' not in wb.sheetnames:
            return {}
        ws = wb['Shifts & Availability']

        year_hint = _date.today().year

        # Row 4 has date headers starting at column C (index 3)
        date_labels: list[str] = []
        date_objects: list[_date | None] = []
        for col_offset in range(14):
            cell = ws.cell(row=4, column=3 + col_offset)
            if not cell.value:
                break
            label = str(cell.value).strip()
            if not any(d in label for d in ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']):
                break
            date_labels.append(label)
            try:
                val = label.split(',', 1)[1].strip() if ',' in label else label
                parsed = _dt.strptime(f"{val}/{year_hint}", "%d/%b/%Y").date()
                date_objects.append(parsed)
            except Exception:
                date_objects.append(None)

        if not date_labels:
            return {}

        result: dict = {d: [] for d in date_objects if d is not None}

        # Driver rows start at row 6, column A
        for row_idx in range(6, (ws.max_row or 200) + 1):
            driver_cell = ws.cell(row=row_idx, column=1)
            if not driver_cell.value:
                break
            driver_name = str(driver_cell.value).strip()
            if 'Total' in driver_name or not driver_name:
                break
            for col_offset, (label, d_obj) in enumerate(zip(date_labels, date_objects)):
                if d_obj is None:
                    continue
                cell = ws.cell(row=row_idx, column=3 + col_offset)
                if cell.value:
                    av = str(cell.value).strip().lower()
                    if av and av != 'unavailable':
                        result[d_obj].append(driver_name)

        return result
    except Exception as exc:
        logger.warning("_parse_all_schedule_dates failed: %s", exc)
        return {}


# ─────────────────────────────────────────────────────────────────────────────
# Classification — filename + user description → detected_type string
# ─────────────────────────────────────────────────────────────────────────────

_TYPE_LABELS = {
    "quality_csv":       "Quality Metrics CSV",
    "dvic":              "DVIC Pre-Trip Under-90s (Excel)",
    "cortex":            "Cortex Routes (Excel)",
    "dop":               "DOP / Dispatch (Excel)",
    "fleet":             "Vehicle Data / Daily Fleet File (Excel)",
    "okami_capacity":    "Okami Capacity Forecast (Next-Day)",
    "driver_schedule":   "Driver Schedule (Excel)",
    "route_sheets":      "Route Sheets (PDF)",
    "wst_zip":           "WST Data (ZIP)",
    "variable_invoice":  "Variable Invoice (PDF)",
    "fleet_invoice":     "Fleet Invoice (PDF)",
    "weekly_incentive":  "Weekly Incentive (PDF)",
    "dsp_scorecard":     "DSP Scorecard (PDF)",
    "pod_report":        "POD Report (PDF)",
    "unknown":           "Unknown — needs manual review",
}


def _classify(filename: str, message: str, channel_id: str = "") -> str:
    name = (filename or "").lower()
    msg = (message or "").lower()
    combined = name + " " + msg
    ext = os.path.splitext(name)[1]
    from_dlv3_info = channel_id == DLV3_INFO_CHANNEL

    if ext == ".zip":
        return "wst_zip"

    if ext == ".csv":
        if any(k in combined for k in ("okami", "capacity forecast", "capacity_forecast", "next day capacity")):
            return "okami_capacity"
        if any(k in combined for k in ("vehicledata", "vehiclesdata", "vehicle data", "vehicles data", "vehicle_data", "fleet", "daily fleet")):
            return "fleet"
        if any(k in combined for k in ("quality", "trailing", "overview", "scorecard")):
            return "quality_csv"
        if re.search(r"\bw\d{2}\b", combined):   # W27, W03, etc.
            return "quality_csv"
        if any(k in combined for k in ("dop", "dispatch ops", "dispatch_ops")):
            return "dop"
        # #dlv3-nday-info only ever carries DOP + route sheet data, so an
        # otherwise-unrecognized CSV dropped there (e.g. a generic "NDAY.csv")
        # is DOP data by convention rather than a real "unknown" file.
        if from_dlv3_info:
            return "dop"
        return "unknown"

    if ext in (".xlsx", ".xls"):
        if any(k in combined for k in ("okami", "capacity forecast", "capacity_forecast", "next day capacity")):
            return "okami_capacity"
        if any(k in combined for k in ("dvic", "pre_trip", "pre-trip", "under90", "under 90", "pretrip")):
            return "dvic"
        if any(k in combined for k in ("cortex", "routes_dlv", "dlv3", "dlv2")):
            return "cortex"
        if any(k in combined for k in ("dop", "dispatch ops", "dispatch_ops")):
            return "dop"
        if any(k in combined for k in ("vehicledata", "vehiclesdata", "vehicle data", "vehicles data", "vehicle_data", "fleet", "daily fleet")):
            return "fleet"
        if any(k in combined for k in ("schedule", "shift", "availability", "rostered work")):
            return "driver_schedule"
        if from_dlv3_info:
            return "dop"
        return "unknown"

    if ext == ".pdf":
        if any(k in combined for k in ("okami", "capacity forecast", "capacity_forecast", "next day capacity")):
            return "okami_capacity"
        if any(k in combined for k in ("variable invoice", "variable_invoice")):
            return "variable_invoice"
        if any(k in combined for k in ("fleet invoice", "fleet_invoice")):
            return "fleet_invoice"
        if any(k in combined for k in ("incentive",)):
            return "weekly_incentive"
        if any(k in combined for k in ("scorecard", "dsp score")):
            return "dsp_scorecard"
        if any(k in combined for k in ("pod report", "pod_report")):
            return "pod_report"
        if any(k in combined for k in ("route sheet", "route_sheet")):
            return "route_sheets"
        # #dlv3-nday-info's only PDF file type is the daily route sheet.
        if from_dlv3_info:
            return "route_sheets"
        return "unknown"

    return "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Channel scan (called by background loop + manual Scan Now button)
# ─────────────────────────────────────────────────────────────────────────────

def scan_ops_channel(db: Session) -> list[str]:
    """
    Scan #nday-operations-management and #dlv3-nday-info for new file shares.
    Creates OpsIngestJob rows for anything not yet seen.
    Returns list of new filenames detected.
    """
    client = _slack_client()
    if not client:
        logger.warning("Ops ingest: SLACK_BOT_TOKEN not set — scan skipped.")
        return []

    known_ids: set[str] = {
        row[0] for row in db.query(OpsIngestJob.slack_file_id).all()
    }

    new_filenames: list[str] = []
    for channel_id in SCAN_CHANNELS:
        try:
            resp = client.conversations_history(channel=channel_id, limit=100)
        except Exception as exc:
            logger.warning("Ops channel history failed for %s: %s", channel_id, exc)
            continue

        for msg in resp.get("messages", []):
            files = msg.get("files", [])
            if not files:
                continue

            msg_text: str = msg.get("text", "") or ""
            msg_ts: str = msg.get("ts", "") or ""

            for f in files:
                fid: str = f.get("id", "")
                if not fid or fid in known_ids:
                    continue

                fname: str = f.get("name", "") or "unknown"
                url: str = f.get("url_private_download") or f.get("url_private") or ""
                dtype = _classify(fname, msg_text, channel_id)

                job = OpsIngestJob(
                    slack_file_id=fid,
                    slack_message_ts=msg_ts,
                    slack_message_text=msg_text[:1000] if msg_text else None,
                    file_name=fname,
                    file_url=url,
                    detected_type=dtype,
                    status="pending",
                    detected_at=datetime.now(timezone.utc),
                )
                db.add(job)
                known_ids.add(fid)
                new_filenames.append(fname)
                logger.info("Ops ingest: queued %s as %s (from %s)", fname, dtype, channel_id)

    if new_filenames:
        db.commit()

    return new_filenames


# ─────────────────────────────────────────────────────────────────────────────
# Ingest dispatch — handlers added here as each file type is introduced
# ─────────────────────────────────────────────────────────────────────────────

def _dispatch(job: OpsIngestJob, content: bytes, db: Session) -> dict:
    t = job.detected_type

    # ── Quality Metrics CSV ──────────────────────────────────────────────────
    if t == "quality_csv":
        from api.src.routes.quality import _store_quality_metrics
        return _store_quality_metrics(content, job.file_name, job.slack_file_id, db)

    # ── Cortex Routes Excel ──────────────────────────────────────────────────
    if t == "cortex":
        ext = os.path.splitext(job.file_name)[1].lower() or ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            from api.src.orchestrator import orchestrator
            from api.src.database import Cortex, ensure_cortex_driver_name_column
            from api.src.routes.uploads import _infer_file_date
            orchestrator.ingest_cortex(tmp_path)
            ensure_cortex_driver_name_column()
            upload_date = _infer_file_date(job.file_name) or datetime.utcnow().date()
            db.query(Cortex).filter(Cortex.source_file == job.file_name).delete(synchronize_session=False)
            for rec in orchestrator.status.cortex_records:
                db.add(Cortex(
                    assignment_date=upload_date,
                    dsp_code=rec.dsp,
                    route_code=rec.route_code,
                    service_type=rec.delivery_service_type,
                    driver_name=getattr(rec, "driver_name", None),
                    source_file=job.file_name,
                    **({ "transporter_id": rec.transporter_id }
                       if hasattr(rec, "transporter_id") else {}),
                ))
            db.commit()

            # Trigger day-of DMs now that route assignments are populated.
            # Each call gets its own try/except — a DM failure must not
            # silently block the assignment matrix (and vice versa).
            from api.src.routes.rostering import send_day_of_dms, post_assignment_matrix
            try:
                send_day_of_dms(upload_date, db)
            except Exception:
                logger.exception("Day-of DM trigger after Cortex ingest failed (upload_date=%s)", upload_date)
            try:
                post_assignment_matrix(upload_date, db)
            except Exception:
                logger.exception("Assignment-matrix trigger after Cortex ingest failed (upload_date=%s)", upload_date)

            return {"status": "ingested", "records": len(orchestrator.status.cortex_records)}
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── DOP Excel ────────────────────────────────────────────────────────────
    if t == "dop":
        ext = os.path.splitext(job.file_name)[1].lower() or ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            from api.src.orchestrator import orchestrator
            from api.src.database import DOP
            from api.src.routes.uploads import _infer_file_date
            orchestrator.ingest_dop(tmp_path)
            upload_date = _infer_file_date(job.file_name) or datetime.utcnow().date()
            db.query(DOP).filter(DOP.source_file == job.file_name).delete(synchronize_session=False)
            for rec in orchestrator.status.dop_records:
                db.add(DOP(
                    schedule_date=upload_date,
                    station=getattr(rec, "staging_location", None),
                    dsp_code=rec.dsp,
                    route_code=rec.route_code,
                    wave=rec.wave,
                    planned_packages=getattr(rec, "num_packages", None),
                    route_duration=getattr(rec, "route_duration", None),
                    service_type=getattr(rec, "service_type", None),
                    source_file=job.file_name,
                ))
            db.commit()
            return {"status": "ingested", "records": len(orchestrator.status.dop_records)}
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── Route Sheets PDF ─────────────────────────────────────────────────────
    if t == "route_sheets":
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            from api.src.orchestrator import orchestrator
            from api.src.database import RouteSheet
            from api.src.routes.uploads import _archive_upload
            orchestrator.ingest_route_sheets([tmp_path])
            records = orchestrator.status.route_sheets
            db.add(RouteSheet(
                upload_date=datetime.utcnow().date(),
                file_name=job.file_name,
                file_size=len(content),
                processing_status="processed",
                total_routes=len(records),
                total_assignments=0,
                processed_at=datetime.utcnow(),
            ))
            _archive_upload(db, upload_type="route_sheets", source_file=job.file_name,
                             payload=records, record_count=len(records))
            db.commit()
            return {"status": "ingested", "records": len(records)}
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── Vehicle Data / Daily Fleet File (Excel or CSV) ──────────────────────
    if t == "fleet":
        ext = os.path.splitext(job.file_name)[1].lower() or ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            from api.src.orchestrator import orchestrator
            from api.src.database import Vehicle
            orchestrator.ingest_fleet(tmp_path)
            records = orchestrator.status.fleet_records
            updated = created = 0
            for record in records:
                existing = db.query(Vehicle).filter(Vehicle.vin == record.vin).first()
                status_value = str(record.operational_status or "active").lower()
                if existing:
                    existing.vehicle_name = record.vehicle_name
                    existing.service_type = record.service_type
                    existing.status = status_value
                    updated += 1
                else:
                    db.add(Vehicle(
                        vin=record.vin,
                        vehicle_name=record.vehicle_name,
                        service_type=record.service_type,
                        status=status_value,
                    ))
                    created += 1
            db.commit()
            return {"status": "ingested", "records": len(records), "created": created, "updated": updated}
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── Driver Schedule Excel ────────────────────────────────────────────────
    if t == "driver_schedule":
        ext = os.path.splitext(job.file_name)[1].lower() or ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name
        try:
            from api.src.orchestrator import orchestrator
            from api.src.database import DriverScheduleEntry
            ok = orchestrator.ingest_driver_schedule(tmp_path)
            if not ok:
                errs = orchestrator.status.validation_errors[-3:]
                return {"status": "error", "message": "; ".join(errs) or "Parse failed"}
            orchestrator.generate_driver_schedule_report(compact=True)

            summary = orchestrator.status.driver_schedule
            saved_count = 0
            dates_saved = []
            try:
                # Parse ALL dates from the Shifts & Availability tab so the
                # callout page can filter by any shift date in the week.
                by_date = _parse_all_schedule_dates(tmp_path)
                if by_date:
                    all_dates = list(by_date.keys())
                    db.query(DriverScheduleEntry).filter(
                        DriverScheduleEntry.schedule_date.in_(all_dates)
                    ).delete(synchronize_session=False)
                    for sched_date, drivers in by_date.items():
                        for driver_name in drivers:
                            db.add(DriverScheduleEntry(
                                schedule_date=sched_date,
                                driver_name=driver_name,
                                source_file=job.file_name,
                            ))
                            saved_count += 1
                    # Annotate wave/show times for the primary scheduled date
                    if summary and summary.date:
                        from datetime import datetime as _dt
                        try:
                            primary_date = _dt.strptime(summary.date, "%m/%d/%Y").date()
                            sweeper_names = set(summary.sweepers or [])
                            assign_map = {a.driver_name: a for a in (summary.assignments or [])}
                            entries = (
                                db.query(DriverScheduleEntry)
                                .filter(DriverScheduleEntry.schedule_date == primary_date)
                                .all()
                            )
                            for entry in entries:
                                if entry.driver_name in sweeper_names:
                                    entry.is_sweeper = True
                                a = assign_map.get(entry.driver_name)
                                if a:
                                    entry.wave_time = getattr(a, 'wave_time', None)
                                    entry.show_time = getattr(a, 'show_time', None)
                                    entry.service_type = getattr(a, 'service_type', None)
                        except Exception as e:
                            logger.warning("Schedule annotation failed: %s", e)
                    # Upsert DriverRosterEntry for every unique driver in the
                    # schedule so the callout page PIN check can find them.
                    from api.src.database import DriverRosterEntry as DRE
                    all_names = {n for dl in by_date.values() for n in dl}
                    existing_names = {
                        r.payroll_name
                        for r in db.query(DRE.payroll_name)
                            .filter(DRE.payroll_name.in_(list(all_names)))
                            .all()
                    }
                    for name in all_names - existing_names:
                        db.add(DRE(payroll_name=name, is_active=True, ssn_last4="1234"))
                    db.commit()
                    dates_saved = [d.isoformat() for d in all_dates]
            except Exception as e:
                logger.warning("Schedule DB persist failed: %s", e)

            # Trigger driver shift DMs + #nday-mgt summary for tomorrow's date
            primary_date_obj = None
            if summary and summary.date:
                try:
                    from datetime import datetime as _dt2
                    primary_date_obj = _dt2.strptime(summary.date, "%m/%d/%Y").date()
                except Exception:
                    pass
            if primary_date_obj:
                try:
                    from api.src.routes.rostering import send_driver_shift_dms, post_mgt_summary
                    send_driver_shift_dms(primary_date_obj, db)
                    post_mgt_summary(primary_date_obj, db)
                except Exception as e:
                    logger.warning("Post-schedule DM/summary trigger failed: %s", e)

            return {
                "status": "ingested",
                "schedule_date": summary.date if summary else None,
                "drivers_saved": saved_count,
                "dates_saved": dates_saved,
                "report_path": orchestrator.status.driver_schedule_report_path,
            }
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # ── DVIC Pre-Trip Under-90s Excel ────────────────────────────────────────
    if t == "dvic":
        from api.src.routes.dvic import _store_dvic
        return _store_dvic(content, job.file_name, job.slack_file_id, db)

    # ── DSP Scorecard PDF ────────────────────────────────────────────────────
    if t == "dsp_scorecard":
        from api.src.routes.dsp_scorecard_weekly import _store_scorecard, _post_summary_to_slack
        snap = _store_scorecard(content, job.file_name, job.slack_file_id, db)
        _post_summary_to_slack(snap)
        return {"status": "ingested", "week": snap.week, "overall": snap.overall_standing}

    # ── More handlers added here as file types are introduced ────────────────

    return {"status": "unsupported", "message": f"No ingest handler yet for type '{t}'. Will be added when this file type is introduced."}


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────

def _job_to_dict(j: OpsIngestJob) -> dict:
    return {
        "id": j.id,
        "slack_file_id": j.slack_file_id,
        "file_name": j.file_name,
        "detected_type": j.detected_type,
        "type_label": _TYPE_LABELS.get(j.detected_type, j.detected_type),
        "description": j.slack_message_text,
        "status": j.status,
        "result": j.result_json,
        "error_message": j.error_message,
        "detected_at": j.detected_at.isoformat() if j.detected_at else None,
        "ingested_at": j.ingested_at.isoformat() if j.ingested_at else None,
    }


@router.get("/jobs")
def list_jobs(status: Optional[str] = None, limit: int = 100, db: Session = Depends(get_db)):
    """Return ops ingest jobs. Pass ?status=pending|complete|error|skipped to filter."""
    q = db.query(OpsIngestJob).order_by(OpsIngestJob.detected_at.desc())
    if status:
        q = q.filter(OpsIngestJob.status == status)
    jobs = q.limit(limit).all()
    pending = sum(1 for j in jobs if j.status == "pending")
    return {"total": len(jobs), "pending": pending, "jobs": [_job_to_dict(j) for j in jobs]}


@router.post("/scan")
def manual_scan(db: Session = Depends(get_db)):
    """Trigger an immediate scan of #nday-operations-management and #dlv3-nday-info for new files."""
    new_files = scan_ops_channel(db)
    return {
        "status": "scanned",
        "new_files_detected": len(new_files),
        "filenames": new_files,
    }


@router.post("/jobs/{job_id}/ingest")
def ingest_job(job_id: int, db: Session = Depends(get_db)):
    """Download the Slack file and run it through the appropriate ingest handler."""
    job = db.query(OpsIngestJob).filter(OpsIngestJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found.")
    if job.status not in ("pending", "error"):
        raise HTTPException(400, f"Job status is '{job.status}' — only pending or error jobs can be ingested.")
    if not job.file_url:
        raise HTTPException(400, "No download URL stored for this job.")

    job.status = "ingesting"
    db.commit()

    content = _download_file(job.file_url)
    if not content:
        job.status = "error"
        job.error_message = "Could not download file from Slack. The link may have expired."
        db.commit()
        raise HTTPException(502, job.error_message)

    try:
        result = _dispatch(job, content, db)
    except Exception as exc:
        logger.exception("Ops ingest dispatch error for job %s", job_id)
        job.status = "error"
        job.error_message = str(exc)[:500]
        job.result_json = None
        db.commit()
        raise HTTPException(500, f"Ingest failed: {exc}")

    job.status = "complete" if result.get("status") not in ("error", "unsupported") else result["status"]
    job.result_json = result
    job.ingested_at = datetime.now(timezone.utc)
    db.commit()

    # Post Slack confirmation
    label = _TYPE_LABELS.get(job.detected_type, job.detected_type)
    if result.get("status") == "ingested":
        records = result.get("records") or result.get("driver_count") or result.get("records_parsed", "")
        confirm_text = f":white_check_mark: *{label} ingested* — `{job.file_name}`"
        if records:
            confirm_text += f"\nRecords loaded: {records}"
        if result.get("week"):
            confirm_text += f" | Week: {result['week']}"
        _post_confirmation(confirm_text)
    elif result.get("status") == "already_ingested":
        _post_confirmation(f":repeat: `{job.file_name}` was already ingested (skipped duplicate).")
    elif result.get("status") == "unsupported":
        _post_confirmation(f":warning: `{job.file_name}` queued but no handler built yet for type `{job.detected_type}`.")

    return _job_to_dict(job)


@router.post("/jobs/{job_id}/skip")
def skip_job(job_id: int, db: Session = Depends(get_db)):
    """Mark a job as skipped (won't be ingested)."""
    job = db.query(OpsIngestJob).filter(OpsIngestJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found.")
    job.status = "skipped"
    db.commit()
    return _job_to_dict(job)


@router.patch("/jobs/{job_id}/type")
def reclassify_job(job_id: int, detected_type: str, db: Session = Depends(get_db)):
    """Manually override the detected_type for a job before ingesting."""
    if detected_type not in _TYPE_LABELS:
        raise HTTPException(400, f"Unknown type. Valid: {list(_TYPE_LABELS.keys())}")
    job = db.query(OpsIngestJob).filter(OpsIngestJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found.")
    job.detected_type = detected_type
    job.status = "pending"
    db.commit()
    return _job_to_dict(job)
