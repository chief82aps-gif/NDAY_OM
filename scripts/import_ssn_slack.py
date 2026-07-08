"""
Import real SSN last-4 PINs and Slack user IDs into driver_roster.

Usage:
    python scripts/import_ssn_slack.py \
        --ssn   "C:/Users/chief/Downloads/SSN (1).xlsx" \
        --slack "C:/Users/chief/Downloads/nday-team-room_slack_members (1).xlsx" \
        --db-url "postgresql://..." \
        [--dry-run]

Without --db-url uses DATABASE_URL env var (or local SQLite dev.db).
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

# ── path setup ────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent.parent))


def _norm(s: str) -> str:
    """Lowercase, collapse whitespace, strip punctuation for fuzzy compare."""
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9 ]", "", s.lower())).strip()


def _ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


def _build_ssn_candidates(first: str, middle: str, last: str) -> list[str]:
    """All plausible full-name formats from SSN file columns."""
    first  = (first  or "").strip()
    middle = (middle or "").strip()
    last   = (last   or "").strip()
    candidates = []
    if first and middle and last:
        candidates.append(f"{first} {middle} {last}")
        candidates.append(f"{first} {last}")
        candidates.append(f"{last} {first} {middle}")
        candidates.append(f"{last} {first}")
    elif first and last:
        candidates.append(f"{first} {last}")
        candidates.append(f"{last} {first}")
    return candidates


def load_ssn(path: str) -> list[dict]:
    """Load SSN xlsx → list of {candidates: [...], last4}"""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdrs = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(hdrs, raw))
        last4 = str(row.get("last 4") or "").strip().zfill(4)
        first  = str(row.get("Legal First Name")   or "").strip()
        middle = str(row.get("Legal Middle Name")  or "").strip()
        last   = str(row.get("Legal Last Name")    or "").strip()
        if not last4 or not (first or last):
            continue
        rows.append({"candidates": _build_ssn_candidates(first, middle, last), "last4": last4})
    return rows


def load_slack(path: str) -> list[dict]:
    """Load Slack xlsx → list of {user_id, username, display_name}"""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    hdrs = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(hdrs, raw))
        user_id      = str(row.get("User ID")      or "").strip()
        username     = str(row.get("Username")      or "").strip()
        display_name = str(row.get("Display Name") or "").strip()
        if user_id:
            rows.append({"user_id": user_id, "username": username, "display_name": display_name})
    return rows


MATCH_THRESHOLD = 0.82   # SequenceMatcher ratio cutoff


def best_ssn_match(roster_name: str, ssn_rows: list[dict]) -> tuple[str | None, float]:
    """Find the SSN row whose candidate names best match roster_name."""
    best_last4, best_score = None, 0.0
    for row in ssn_rows:
        for cand in row["candidates"]:
            score = _ratio(roster_name, cand)
            if score > best_score:
                best_score = score
                best_last4 = row["last4"]
    if best_score >= MATCH_THRESHOLD:
        return best_last4, best_score
    return None, best_score


def best_slack_match(roster_name: str, slack_rows: list[dict]) -> tuple[str | None, str | None, float]:
    """Find Slack user whose display_name or username best matches roster_name."""
    best_id, best_display, best_score = None, None, 0.0
    parts = _norm(roster_name).split()
    first = parts[0] if parts else ""
    last  = parts[-1] if len(parts) > 1 else ""

    for row in slack_rows:
        candidates = []
        if row["display_name"]:
            candidates.append(row["display_name"])
        if row["username"]:
            candidates.append(row["username"])
        for cand in candidates:
            score = _ratio(roster_name, cand)
            # Boost if first OR last name appears in the candidate
            nc = _norm(cand)
            if first in nc or last in nc:
                score = max(score, 0.60)
            if score > best_score:
                best_score = score
                best_id      = row["user_id"]
                best_display = row["display_name"] or row["username"]

    if best_score >= MATCH_THRESHOLD:
        return best_id, best_display, best_score
    return None, None, best_score


def run(ssn_path: str, slack_path: str, db_url: str | None, dry_run: bool):
    db_url = db_url or os.getenv("DATABASE_URL", "sqlite:///./dev.db")
    os.environ["DATABASE_URL"] = db_url

    from api.src.database import SessionLocal, DriverRosterEntry, Base, engine
    Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    roster = db.query(DriverRosterEntry).all()
    print(f"\nRoster loaded: {len(roster)} entries")

    ssn_rows   = load_ssn(ssn_path)   if ssn_path   else []
    slack_rows = load_slack(slack_path) if slack_path else []
    print(f"SSN file:   {len(ssn_rows)} employees")
    print(f"Slack file: {len(slack_rows)} members\n")

    ssn_hits = ssn_misses = 0
    slack_hits = slack_misses = 0

    for entry in roster:
        name = entry.payroll_name

        # ── SSN match ────────────────────────────────────────────────────────
        if ssn_rows:
            last4, score = best_ssn_match(name, ssn_rows)
            if last4:
                print(f"  SSN  [{score:.2f}] {name!r:45s}  → PIN {last4}")
                if not dry_run:
                    entry.ssn_last4 = last4
                ssn_hits += 1
            else:
                print(f"  SSN  [NO MATCH {score:.2f}] {name!r}")
                ssn_misses += 1

        # ── Slack match ───────────────────────────────────────────────────────
        if slack_rows:
            uid, display, score = best_slack_match(name, slack_rows)
            if uid:
                print(f"  Slack[{score:.2f}] {name!r:45s}  → {uid} ({display})")
                if not dry_run:
                    entry.slack_member_id    = uid
                    entry.slack_display_name = display
                slack_hits += 1
            else:
                # Only log Slack misses at a lower verbosity (most drivers won't be in channel)
                slack_misses += 1

    if not dry_run:
        db.commit()
        print(f"\n✅  Committed to database.")
    else:
        print(f"\n⚠️  DRY RUN — no changes written.")

    db.close()
    print(f"\nSSN:   {ssn_hits} matched / {ssn_misses} unmatched")
    print(f"Slack: {slack_hits} matched / {slack_misses} unmatched (most drivers not in channel — expected)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import SSN last-4 PINs and Slack IDs into driver_roster")
    parser.add_argument("--ssn",   default="C:/Users/chief/Downloads/SSN (1).xlsx",
                        help="Path to SSN xlsx file")
    parser.add_argument("--slack", default="C:/Users/chief/Downloads/nday-team-room_slack_members (1).xlsx",
                        help="Path to Slack members xlsx file")
    parser.add_argument("--db-url", default=None,
                        help="Database URL (overrides DATABASE_URL env var)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview matches without writing to DB")
    args = parser.parse_args()
    run(args.ssn, args.slack, args.db_url, args.dry_run)
