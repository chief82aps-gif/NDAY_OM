"""Regression test for driver schedule date alignment.

Ensures the parsed scheduled date and assignment rows align to the same date
from the uploaded workbook (not the current system date).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl

from api.src.ingest.driver_schedule import parse_driver_schedule_excel


def _parse_timestamp_date(timestamp_value: str) -> datetime.date:
    candidates = [
        "%m/%d/%y, %I:%M:%S %p",
        "%m/%d/%Y, %I:%M:%S %p",
        "%m/%d/%y %I:%M:%S %p",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%y",
        "%m/%d/%Y",
    ]
    raw = str(timestamp_value).strip()
    for fmt in candidates:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise AssertionError(f"Unable to parse timestamp date from A2: {raw}")


def _parse_header_date(header_value: str, year: int) -> Optional[datetime.date]:
    raw = str(header_value).strip()
    try:
        if "," in raw:
            raw = raw.split(",", 1)[1].strip()
        parsed = datetime.strptime(f"{raw}/{year}", "%d/%b/%Y")
        return parsed.date()
    except ValueError:
        return None


def run_regression_test() -> None:
    workbook_path = Path("uploads/driver_schedule_Test_Schedule.xlsx")
    assert workbook_path.exists(), f"Missing fixture: {workbook_path}"

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb["Rostered Work Blocks"]

    timestamp_date = _parse_timestamp_date(ws["A2"].value)

    expected_label = None
    for col in range(3, 13):
        label = ws.cell(row=4, column=col).value
        if not label:
            continue
        parsed = _parse_header_date(label, timestamp_date.year)
        if parsed == timestamp_date:
            expected_label = str(label).strip()
            break

    assert expected_label, "No date header matched timestamp date"

    summary, errors = parse_driver_schedule_excel(str(workbook_path))

    assert not errors, f"Unexpected parser errors: {errors}"
    assert summary.date == timestamp_date.strftime("%m/%d/%Y"), (
        f"Scheduled date mismatch: expected {timestamp_date:%m/%d/%Y}, got {summary.date}"
    )
    assert summary.assignments, "No assignments parsed"

    mismatched = [a.driver_name for a in summary.assignments if a.date != expected_label]
    assert not mismatched, (
        "Assignments include rows from non-selected dates; sample mismatches: "
        + ", ".join(mismatched[:5])
    )

    print("PASS: driver schedule date alignment regression test")
    print(f"  scheduled_date: {summary.date}")
    print(f"  assignment_date_label: {expected_label}")
    print(f"  assignments_count: {len(summary.assignments)}")


if __name__ == "__main__":
    run_regression_test()
